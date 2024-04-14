import argparse
import csv
import datetime
import functools
import multiprocessing
import os
import re
import sys
import tempfile
import time
import zipfile
from itertools import chain, zip_longest
from signal import SIGINT, SIG_IGN, signal

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import Process

version = "2.0"
output_name = datetime.datetime.now().strftime("Cryptoscan_%Y_%m_%d_%H%M%S.csv")


class StatsTracker:
    def __init__(self):
        self.processed_files_count = 0
        self.total_bytes_processed = 0
        self.count_seedstrings = 0
        self.count_addresses = 0


def init_worker():
    signal(SIGINT, SIG_IGN)


def convertsizestring_to_bytesint(size_str):
    match = re.match(r'(\d+)\s*([KMG]?B)', size_str, re.IGNORECASE)
    if not match:
        raise ValueError("Invalid size format")

    size, unit = match.groups()
    size = int(size)

    unit = unit.upper()
    unit_factors = {'B': 1, 'KB': 1024, 'MB': 1024**2, 'GB': 1024**3}

    if unit not in unit_factors:
        raise ValueError("Invalid unit of size")

    return int(size * unit_factors[unit]), size_str


def convertbytesint_to_sizestring(size_int):
    for unit in ['B', 'KB', 'MB', 'GB']:
        if size_int < 1024.0:
            break
        size_int /= 1024.0
    return f"{size_int:.{0}f}{unit}"


def process_result(result, lock, file, stats_):
    with lock:
        try:
            if result[0]:  # Special check for result from too small file in zipfile
                if isinstance(result[0][0][0], tuple):  # Unpacks result if it's a result from an archive
                    for single_result in result[0]:
                        write_to_csv(single_result, file, stats_)
                else:
                    write_to_csv(result, file, stats_)
        except Exception as err:
            print("Error in processing result: " + str(err))


def write_to_csv(result, file, stats_):
    try:
        for pattern, address, offset in zip_longest(result[0][0], result[0][1], result[0][2], fillvalue='N/A'):
            if "BIP-39 Seed String" in pattern:
                stats_.count_seedstrings += 1
            else:
                stats_.count_addresses += 1
            file.write(str(pattern) + "," + str(address) + "," + str(result[1]) + "," + str(offset) + '\n')
            file.flush()
    except Exception as err:
        print("Error in writing to .csv: " + str(err))


def process_single_file(path, max_filesize, excl_paths, stats_, temppath_):
    lock = multiprocessing.Lock()
    extension = os.path.splitext(path)[1].lower()
    if extension == '.ufdr' or extension == '.zip':
        process_single_archive(path, max_filesize, excl_paths, stats_, temppath_)
    else:
        result = Process.process_file(1000000000000, excl_paths, None, temppath_, path)
        try:
            stats_.processed_files_count += 1
            stats_.total_bytes_processed += int(result[2])
            if result and all(result[0]):
                with open(output_name, 'a') as file:
                    file.write('Pattern,Found addresses,Filename,Offset\n')
                    process_result(result, lock, file, stats_)
        except Exception as err:
            print(f'Error: {err}')


def process_single_archive(path, max_filesize, excl_paths, stats_, temppath_):
    if temppath_:
        temp_dir = tempfile.TemporaryDirectory(dir=temppath_)
    else:
        temp_dir = tempfile.TemporaryDirectory()
    with temp_dir as temp_dir:
        with zipfile.ZipFile(path, 'r') as archive_ref:
            print("Processing single archive file. Make sure you have enough storage space as all files will be extracted to a temp directory.")
            archive_ref.extractall(temp_dir)
            cpucount = multiprocessing.cpu_count() - 2
            pool = multiprocessing.Pool(cpucount, init_worker)
            lock = multiprocessing.Lock()
            file_generator = chain.from_iterable((os.path.join(root, file) for file in files) for root, dirs, files in os.walk(temp_dir))

            try:
                with open(output_name, 'a') as file:
                    file.write('Pattern,Found addresses,Filename,Offset\n')
                    run_worker = pool.imap_unordered(functools.partial(Process.process_file, max_filesize, excl_paths, None, temppath_), file_generator)
                    for result in run_worker:
                        if result:
                            stats_.processed_files_count += 1
                            stats_.total_bytes_processed += int(result[2])
                            if all(result[0]):
                                process_result(result, lock, file, stats_)

            except Exception as err:
                print(f'Error: {err}')
            pool.close()
            pool.join()


def process_directory(path, max_filesize, excl_paths, stats_, temppath_):
    cpucount = multiprocessing.cpu_count() - 2
    pool = multiprocessing.Pool(cpucount, init_worker)
    lock = multiprocessing.Lock()
    file_generator = chain.from_iterable((os.path.join(root, file) for file in files) for root, dirs, files in os.walk(path))

    try:
        with open(output_name, 'a') as file:
            file.write('Pattern,Found addresses,Filename,Offset\n')
            run_worker = pool.imap_unordered(functools.partial(Process.process_file, max_filesize, excl_paths, None, temppath_), file_generator)
            for result in run_worker:
                if result:
                    stats_.processed_files_count += 1
                    stats_.total_bytes_processed += int(result[2])
                    if all(result[0]):
                        process_result(result, lock, file, stats_)

    except Exception as err:
        print(f'Error: {err}')

    pool.close()
    pool.join()


def startprocessing(search_path, max_filesize, excluded_paths, stats, temp_path):
    if os.path.isfile(search_path):
        process_single_file(search_path, max_filesize, excluded_paths, stats, temp_path)
    elif os.path.isdir(search_path):
        process_directory(search_path, max_filesize, excluded_paths, stats, temp_path)


def usage_and_arguments():
    tool_name = f"""--- Cryptoscan {version} ---"""
    tool_description = f"""This tool recursively scans a specified file or path for files that might contain 
cryptocurrency addresses or seed strings. Additionally, it searches for common wallet file names and paths. 
    """

    print(tool_name)

    parser = argparse.ArgumentParser(allow_abbrev=False, formatter_class=argparse.RawTextHelpFormatter)

    parser.add_argument('path', type=str, help='The path or file to search in.')
    parser.add_argument('--maxfilesize', type=convertsizestring_to_bytesint, default='20GB', help='Optional: Maximum file size to scan (e.g. 10B, 10KB, 10MB, 10GB). Default is 20GB.')
    parser.add_argument('--excludepaths', type=str, nargs='*', default=[], help='Optional: A list of directories to exclude from the search.')
    parser.add_argument('--temppath', type=str, help='Optional: Set a specific temporary directory path.')
    parser.add_argument('--xlsx', action='store_true', help='Optional: Convert the CSV output to an Excel file.\n ')

    if len(sys.argv) <= 1:
        print(tool_description)
        parser.print_help()
        sys.exit(0)

    args = parser.parse_args()

    if not os.path.exists(args.path):
        print(f"Error: The specified path '{args.path}' does not exist.")
        parser.print_help()
        sys.exit(1)

    print("Arguments received:")
    print(f"Path: {args.path}")
    print(f"Max file size: {args.maxfilesize[1]} ({args.maxfilesize[0]} bytes)")
    if args.excludepaths:
        print(f"Excluded directories: {args.excludepaths}")
    if args.temppath:
        print(f"Set temporary directory: {args.temppath}")
    if args.xlsx:
        print("CSV output will be converted to Excel format.")
    print()
    return args.path, args.maxfilesize[0], args.excludepaths, args.temppath, args.xlsx


def convert_csv_to_excel(csv_filename):
    wb = Workbook()
    ws = wb.active

    max_length_per_column = {}
    try:
        with open(csv_filename, 'r', encoding='utf-8') as file:
            reader = csv.reader(file)
            for row in reader:
                ws.append(row)
                for idx, cell in enumerate(row):
                    max_length = max_length_per_column.get(idx, 0)
                    max_length_per_column[idx] = max(len(cell), max_length)

        for idx, max_length in max_length_per_column.items():
            adjusted_width = min(max_length + 2, 70)
            ws.column_dimensions[get_column_letter(idx+1)].width = adjusted_width

        wb.save(csv_filename[:-4] + '.xlsx')
    except Exception as err:
        print(f"Error with xlsx convertion: {err}")


if __name__ == '__main__':
    multiprocessing.freeze_support()

    statistics = StatsTracker()

    arguments = usage_and_arguments()

    searchpath, maxfilesize, excludedpaths, temppath, xlsx_check = arguments

    starttime = time.perf_counter()

    startprocessing(searchpath, maxfilesize, excludedpaths, statistics, temppath)

    print()
    print(f"Processing took: {str(datetime.timedelta(seconds=int(time.perf_counter() - starttime)))} processing time")
    print(f"Addresses found: {statistics.count_addresses}")
    print(f"Seed strings found: {statistics.count_seedstrings}")
    print(f"Processed {statistics.processed_files_count} ({convertbytesint_to_sizestring(statistics.total_bytes_processed)}) non-excluded files.")
    print()
    print(f"Saved output to: {output_name}")

    if xlsx_check:
        convert_csv_to_excel(output_name)
        print(f"Converted output file to: {output_name[:-4]}.xlsx")
